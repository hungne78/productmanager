from ilmarebot.common.config import access_key
import hashlib
import hmac
import base64
import requests
import time
import re

from datetime import datetime, timedelta
import os
import sys
import openpyxl
from ilmarebot.common.config import secret_key, uri, url


def get_todays_reservation_details():
    """Retrieve today's reservations and accurately split name & phone number."""
    
    # Get today's date
    today = datetime.today().strftime('%Y-%m-%d')
    year_month = datetime.today().strftime('%Y-%m')
    
    # Define the Excel file path
    if getattr(sys, 'frozen', False):
        BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    
    folder_path = os.path.join(BASE_DIR, 'excel')
    file_name = f"{year_month}.xlsx"
    file_path = os.path.join(folder_path, file_name)
    
    if not os.path.exists(file_path):
        return f"Error: File '{file_name}' not found in '{folder_path}'."

    try:
        # Load workbook and select today's sheet
        workbook = openpyxl.load_workbook(file_path, read_only=True)
        
        if today not in workbook.sheetnames:
            return f"Error: Sheet '{today}' not found in '{file_name}'."
        
        sheet = workbook[today]
        
        # Retrieve only the 3rd column (Reservation values)
        reservation_details = []
        for row in sheet.iter_rows(min_row=2, values_only=True):  # Skipping header row
            if len(row) >= 3 and row[2]:  # Ensure the 3rd column exists and is not empty
                name, phone_number = split_name_phone(row[2])
                reservation_details.append((name, phone_number))
        
        
        if not reservation_details:
            return f"No reservations found for today ({today})."
        
        return reservation_details

    except Exception as e:
        return f"Error processing reservations: {e}"
    

def split_name_phone(reservation_str):
    # 010-XXXX-XXXX 또는 050X-XXXX-XXXX 형식의 전화번호를 찾는 정규식
    phone_pattern = re.compile(r"(010-\d{4}-\d{4}|050\d-\d{4}-\d{4})")
    match = phone_pattern.search(reservation_str)
    
    if match:
        phone_number = match.group()  # 전화번호 추출
        name = reservation_str[:match.start()].strip()  # 전화번호 앞의 부분만 이름으로 저장
    else:
        name = reservation_str.strip()  # 전화번호가 없을 경우 전체를 이름으로 저장
        phone_number = "Unknown"

    return name, phone_number



def tide_information_process(reference_date=None):
    target_date = datetime.today().strftime('%Y-%m-%d')

    try:
        if getattr(sys, 'frozen', False):
            BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
        else:
            BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(BASE_DIR, 'excel')  # 'excel' 폴더 경로 설정
        file_path = os.path.join(folder_path, 'sea_road.xlsx')  # 'excel' 폴더 내의 파일 경로 설정

        # Load the Excel file
        if not os.path.exists(file_path):
            print("File not found. Please check the file path.")
            return None


        # Load the Excel file
        workbook = openpyxl.load_workbook(file_path)
        # ui.system_status.append(f"{file_path}이 정상적으로 열렸렸습니다.")
        sheet = workbook.active
    except FileNotFoundError:
        print("File not found. Please check the file path.")
        return None
    
    # 날짜 계산: 기준 날짜를 고정
    if reference_date is None:
        reference_date = datetime.now().date()  # 기준 날짜를 고정
        
    if target_date.lower() == "오늘":
        target_date = reference_date
    elif target_date.lower() == "내일":
        target_date = reference_date + timedelta(days=1)
    elif target_date.lower() == "어제":
        target_date = reference_date - timedelta(days=1)
    elif target_date.lower() == "내일 모레" or target_date.lower() == "내일모레" or target_date.lower() == "모레":
        target_date = reference_date + timedelta(days=2)
    
    else:
        # Parse target_date from string
        try:
            target_date = datetime.strptime(target_date, '%Y-%m-%d').date()
        except ValueError:
            return "Error: 날짜 형식이 잘못되었습니다. 'YYYY-MM-DD' 형식을 사용하세요."
    
    # Search for the target date and fetch 4 values
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, datetime) and cell.value.date() == target_date:
                values = []
                # Start from the 3rd column after the date
                col_index = cell.col_idx + 3
                for i in range(4):
                    next_cell = sheet.cell(row=cell.row, column=col_index + i)
                    value = str(next_cell.value).zfill(4)  # Zero-pad to 4 digits
                    formatted_value = value[:2] + ':' + value[2:]  # Insert ':'
                    values.append(formatted_value)
                
                return f'''제부도 일마레 펜션에서 알려드립니다.
{target_date} 의 제부도 통행가능한 시간은
1차 통행가능시간 : {values[0]}부터 {values[1]}까지
2차 통행가능시간 : {values[2]}부터 {values[3]}까지 입니다.

*주의사항*
기상상태에 따라 바다갈라짐 발생 및 지속시간은 30분 이상 편차가 발생할 수 있습니다.
참고로 {values[1]}부터 {values[2]}까지는 바닷길 통행이 불가능합니다.
*참고사항*
일마레 마트에서 체크인 절차를 도와드립니다. 
제부도 일마레펜션에서는 대형마트, 삼겹살식당, 노래연습장이 있어, 편안하고 즐거운 여행을 위한 다양한 편의시설을 제공합니다.
'''
    
    return None



def make_signature(timestamp):
    global secret_key
    global access_key
    global url

    
    # secret_key는 이미 문자열이므로 그대로 사용합니다.
    secret_key_bytes = bytes(secret_key, 'UTF-8')  # secret_key 변수 이름을 변경하여 문제를 방지합니다.
    
    method = "POST"
    message = method + " " + uri + "\n" + timestamp + "\n" + access_key
    message = bytes(message, 'UTF-8')
    
    signingKey = base64.b64encode(hmac.new(secret_key_bytes, message, digestmod=hashlib.sha256).digest())
    return signingKey


def send_message_via_sens(phone_number, message):
    
    timestamp = str(int(time.time() * 1000))
    """
    Send an SMS message via NAVER SENS.

    Args:
        phone_number (str): The phone number to send the message to.
        message (str): The message to send.
    """
    url = "https://sens.apigw.ntruss.com/sms/v2/services/ncp:sms:kr:327347357522:ddnayo_sms/messages"  # Change this to the actual SENS API endpoint
    headers = {
        "Content-Type": "application/json; charset=utf-8",
        "x-ncp-apigw-timestamp": timestamp, 
        "x-ncp-iam-access-key": access_key,
        "x-ncp-apigw-signature-v2": make_signature(timestamp)
    }
    body = {
        "type": "LMS",
        "from": "01062900006",  # Your registered phone number
        "content": message,
        "messages": [{"to": phone_number}]
    }

    response = requests.post(url, json=body, headers=headers)
    if response.status_code == 202:
        print(f"Message sent successfully to {phone_number}")
    else:
        print(f"Failed to send message. Response: {response.text}")



def tide_message(ui):
    result = get_todays_reservation_details()
    
    # 에러 문자열인지, 정상적인 리스트인지 구분
    if isinstance(result, str):
        # 에러나 메시지 출력
        ui.system_status.append(result)
    else:
        # 예약 리스트라면 반복문 돌면서 각각 꺼내쓰기
        for name, phone_number in result:
            ui.append_chat_status(f"{name}님에게 전화번호: {phone_number}로 물때 메세지를 보냈습니다.")
    
    for na, ph in result:
        message = f"{na}님 "+ tide_information_process()
        clean_number = re.sub(r"-", "", ph)
        send_message_via_sens(clean_number, message)
        
        # print(f"Message sent to {ph}: {message}")
    send_message_via_sens("01062900006", "당일 손님 물때시간표 보내기 완료")


   

def send_to_one_tide(ui):
    message = tide_information_process()
    customer = ui.tide_input.text()
    send_message_via_sens(customer, message)