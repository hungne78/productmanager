from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from selenium.webdriver.common.action_chains import ActionChains
from selenium.common.exceptions import NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException
from bs4 import BeautifulSoup
import pandas as pd
import re
import os
import sys
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side, Alignment
from datetime import datetime, timedelta
import openpyxl

scraping_driver = None #떠나요 스크래핑 드라이버
scraping_data = []  # 스크래핑한 데이터 저장 리스트

def process_reservation_data(file_name: str, ui):   #A function that receives a file name and transfers data from the Excel file of that name to the online ledger.
    # Function to clean and validate date strings
    room_names = ui.load_rooms()
    if getattr(sys, 'frozen', False):
        BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(BASE_DIR, 'excel')

    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    
    if not room_names:
        room_names = ["A201", "A202", "B501", "B502", "B503", "B601", "B602", "B603", "사랑채", "C02", "C03", "C04", "C05", "별채"]
        
    # Loop through each unique year-month in the dates
    

    def clean_and_validate_date(date_str):
        
        if pd.isna(date_str) or not isinstance(date_str, str):
            return None, None  # Return a default value or handle as needed
        
        date_str = date_str.replace('~', '-')
        date_range = date_str.split('-')
        
        if len(date_range) == 6:  # For date ranges like 2024-08-11~2024-08-12
            try:
                start_date = datetime.strptime(f"{date_range[0]}-{date_range[1]}-{date_range[2]}", '%Y-%m-%d')
                end_date = datetime.strptime(f"{date_range[3]}-{date_range[4]}-{date_range[5]}", '%Y-%m-%d')
                if start_date > end_date:
                    raise ValueError(f"Start date {start_date} is after end date {end_date}")
                return start_date, end_date
            except ValueError as e:
                print(f"Error parsing date range '{date_str}': {e}")
                return None, None
        elif len(date_range) == 3:  # For single dates
            try:
                single_date = datetime.strptime(date_str.strip(), '%Y-%m-%d')
                return single_date, single_date
            except ValueError as e:
                print(f"Error parsing date '{date_str}': {e}")
                return None, None
        else:
            print(f"Error: Invalid date format '{date_str}'")
            return None, None

    # Load the reservation data
    reservations_df = pd.read_excel(file_name)

    # Extract the room number and other reservation information
    room_numbers = reservations_df.iloc[:, 0]  # A column (index 0)
    reservation_data = reservations_df.iloc[:, 1:]  # All other columns starting from B

    # Clean the date strings and expand to multiple dates if needed
    dates = reservations_df.iloc[:, 3].apply(clean_and_validate_date)
    print(str(dates))

    all_dates = []
    for start_date, end_date in dates:
        if start_date and end_date:
            all_dates.append([start_date + timedelta(days=x) for x in range((end_date - start_date).days + 1)])

    # Flatten the list of dates and expand the DataFrame
    flat_dates = [date for sublist in all_dates for date in sublist]
    expanded_reservations = pd.DataFrame({
        'Room': [room_numbers.iloc[i] for i in range(len(all_dates)) for _ in all_dates[i]],
        'Date': flat_dates,
        **{f'Col_{i+1}': [reservation_data.iloc[j, i] for j in range(len(all_dates)) for _ in all_dates[j]] for i in range(reservation_data.shape[1])}
    })

    
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    # Convert the 'Date' column to datetime
    expanded_reservations['Date'] = pd.to_datetime(expanded_reservations['Date'], errors='coerce')
    
    # Loop through each unique year-month in the dates
    for date in expanded_reservations['Date'].dt.to_period('M').unique():
        output_file_name = f"{date.year}-{str(date.month).zfill(2)}.xlsx"
        # 파일 경로 설정
        output_path = os.path.join(folder_path, output_file_name)
        if not os.path.exists(output_path) or os.path.getsize(output_path) == 0:
            ui.system_status.append(f"파일 {output_file_name} 이 손상되었거나 존재하지 않습니다. 'generate_yearly_excel_files' 함수를 실행합니다.")
            generate_yearly_excel_files(date.year, room_names=room_names)

        if os.path.exists(output_path):
            # Load the existing file
            workbook = openpyxl.load_workbook(output_path)
        else:
            continue  # Skip if the file does not exist

        # Get the corresponding dates for the current year-month
        current_reservations = expanded_reservations[expanded_reservations['Date'].dt.to_period('M') == date]
        
        for _, row in current_reservations.iterrows():
            single_date = row['Date']
            room = row['Room']
            reservation_info = row[2:]

            sheet_name = single_date.strftime('%Y-%m-%d')
            print(f"Checking for sheet with name: {sheet_name}") 
            
            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                print(f"Writing data to sheet: {sheet_name}")  # Add this line
            else:
                print(f"Sheet {sheet_name} does not exist, skipping.") 
                continue  # Skip if the sheet does not exist
            
            # Find the row with the matching room number
            room_found = False
            for sheet_row in reversed(list(sheet.iter_rows(min_row=2, max_row=sheet.max_row))):
                if sheet_row[0].value == room:
                    print(f"Room {room} found in {sheet_name}")
                    
                    # Check if reservation_info is valid before processing it
                    if reservation_info.empty:
                        print(f"Error: reservation_info for room {room} on {single_date} is empty.")
                        continue  # Skip if reservation_info is invalid

                    # Write the reservation info to the matching row
                    if any(keyword in reservation_info.values for keyword in ["취소완료", "예약이취소", "예약취소"]):
                        for col_num in range(1, len(sheet_row)):
                            sheet_row[col_num].value = ""  # Clear the row for cancelled reservation
                    else:
                        for col_num, value in enumerate(reservation_info, start=1):  # Start at column B (index 1)
                            if pd.isna(value):
                                value = ""  # Handle NaN or missing values by converting to empty string
                            print(f"Writing {value} to column {col_num + 1}")  # Add this line to check value being written
                            sheet_row[col_num].value = value
                    
                    room_found = True
                    break

            if not room_found:
                print(f"Error: Room number {room} not found in sheet {sheet_name}")

        # Save the workbook
        workbook.save(output_path)
        print(f"Saved workbook to {output_path}")  # Add this line
        # Ensure the file is properly saved before opening
       


    print("Data transfer completed successfully.")


def initialize_scraping_driver():
    """떠나요 스크래핑 드라이버."""
    
    global scraping_driver
    chrome_options = webdriver.ChromeOptions()
    chrome_options.add_argument("--start-maximized")
    prefs = {"profile.managed_default_content_settings.images": 2}
    chrome_options.add_experimental_option("prefs", prefs)
    chrome_options.add_argument("--headless")
    scraping_driver = webdriver.Chrome(options=chrome_options)    

def ddnayo_whole_res(ui):  #떠나요 웹스크래핑 프로세스 (예약정보 가져오기)
    
    initialize_scraping_driver()
    
    credentials = ui.get_credentials()
    if credentials is None:  # 입력 취소 또는 오류 처리
        ui.system_status.append("[ERROR] Credentials not provided.")
        return

    naver_id, naver_pw, naver_no1, naver_no2, rv_no1, rv_no2, dn_id, dn_pw, dn_no  = credentials  

    ui.system_status.append(f"[INFO] ID: {dn_id}, Password: {'*' * len(dn_pw)}")    

    def clean_room_name(room_name):
        match = re.search(r'C동\s(C\d{2})호', room_name)
        if match:
            return match.group(1)
        if "게스트하우스" in room_name:
            return "사랑채"
        if "D동" in room_name:
            return "별채"
        match = re.search(r'(\d{3})호', room_name)
        if match:
            room_num = match.group(1)
            if room_num.startswith("5") or room_num.startswith("6"):
                return "B" + room_num
            elif room_num.startswith("2"):
                return "A" + room_num
            return room_num
        return room_name

    def clean_date(date_str):
        date_str = re.sub(r'\([^)]*\)', '', date_str)
        date_str = re.sub(r'\d박', '', date_str)
        return date_str.strip()

    def scrape_data(scraping_driver):
        global scraping_data
        
        scraping_data = []

        target_xpath = "/html/body/div[1]/div/div/div[2]/div/div/div[3]/div[2]/div/table"
        next_button_xpath = "/html/body/div[1]/div/div/div[2]/div/div/div[3]/div[2]/div/table/tfoot/tr/td/div/div[3]/button[2]"

        while True:
            
            inner_html = scraping_driver.find_element(By.XPATH, target_xpath).get_attribute('innerHTML')
            soup = BeautifulSoup(inner_html, "lxml")
            rows = soup.select("tbody tr")
            for row in rows:
                columns = row.find_all("td")
                row_data = []
                for idx, column in enumerate(columns):
                    cell_text = column.get_text(strip=True) + " "
                    div_tags = column.find_all('div')
                    for div_tag in div_tags:
                        # Use a regex pattern to match "jssXXX" where XXX is a number between 001 and 999
                        p_tag = div_tag.find('p', class_=re.compile(r"MuiTypography-root jss\d{3} MuiTypography-body1"))
                        span_tag = div_tag.find('span', class_=re.compile(r"MuiTypography-root jss\d{3} MuiTypography-body1"))
                        
                        if p_tag:
                            try:
                                wait = WebDriverWait(scraping_driver, 10)  # 최대 10초까지 대기
                                p_element = wait.until(EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{p_tag.get_text(strip=True)}')]")))
                                ActionChains(scraping_driver).move_to_element(p_element).perform()
                                time.sleep(1)  # 필요한 경우 대기 시간 유지
                                p_text = p_element.text
                                cell_text = cell_text.replace(p_tag.get_text(strip=True), p_text)
                                ui.system_status.append(f"{p_text} recording")
                            except Exception as e:
                                print(f"p 태그 에러: {e} ")

                        if span_tag:
                            try:
                                wait = WebDriverWait(scraping_driver, 10)  # 최대 10초까지 대기
                                span_element = wait.until(EC.presence_of_element_located((By.XPATH, f"//*[contains(text(), '{span_tag.get_text(strip=True)}')]")))
                                ActionChains(scraping_driver).move_to_element(span_element).perform()
                                time.sleep(1)  # 필요한 경우 대기 시간 유지
                                span_text = span_element.text
                                cell_text = cell_text.replace(span_tag.get_text(strip=True), span_text)
                                ui.system_status.append(f"{span_text} recording")
                            except Exception as e:
                                print(f"span 태그 에러: {e} ")
                    if idx == 2:
                        cell_text = clean_room_name(cell_text.strip())
                    if idx == 3:
                        cell_text = clean_date(cell_text.strip())
                    row_data.append(cell_text.strip())
                while len(row_data) < 12:
                    row_data.append("")
                if len(row_data) > 12:
                    row_data = row_data[:12]
                row_data = row_data[2:] + ["", ""]
                scraping_data.append(row_data)
            try:
                next_button = scraping_driver.find_element(By.XPATH, next_button_xpath)
                if 'Mui-disabled' in next_button.get_attribute('class'):
                    break
                scraping_driver.execute_script("arguments[0].scrollIntoView();", next_button)
                time.sleep(1)
                scraping_driver.execute_script("arguments[0].click();", next_button)
                time.sleep(5)
            except (NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException):
                break

        # 데이터프레임 정의 및 저장
        df = pd.DataFrame(scraping_data, columns=['객실', '예약신청', '예약자', '이용기간', '결제액', '정산금액', '결제수단', '예약상태', '인원추가', '수영장옵션', '바베큐', '피크닉세트'])
        df['이용기간'] = df['이용기간'].apply(lambda x: clean_date(x))
        # 'excel' 폴더 경로
        if getattr(sys, 'frozen', False):
            BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
        else:
            BASE_DIR = os.path.dirname(os.path.abspath(__file__))
        folder_path = os.path.join(BASE_DIR, 'excel')

        # 폴더가 없으면 생성
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)

        # 파일 경로 설정
        output_path = os.path.join(folder_path, 'resinfo.xlsx')
        # print(df.head())  # 데이터 확인

        # 엑셀 파일로 저장
        df.to_excel(output_path, index=False)

        process_reservation_data(output_path)

        try:
            if os.path.exists(output_path):
                os.remove(output_path)
                ui.system_status.append(f"파일 {output_path} 삭제 완료.")
            else:
                ui.system_status.append(f"파일 {output_path}이(가) 존재하지 않습니다.")
        except Exception as e:
            ui.system_status.append(f"파일 삭제 중 오류 발생: {e}")

    def start_scraper():
        global  scraping_driver
        
        initialize_scraping_driver()
        login_and_scrape(scraping_driver)
        

    def login_and_scrape(scraping_driver):
        scraping_driver.get("https://partner.ddnayo.com/login")
        time.sleep(5)
        ui.system_status.append("[INFO] ddnayo page loaded.")
        elem = scraping_driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[2]/div[2]/form/div[2]/div/input")
        elem.send_keys(dn_id)
        time.sleep(2)
        elem1 = scraping_driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[2]/div[2]/form/div[4]/div/input")
        elem1.send_keys(dn_pw)
        time.sleep(2)
        elem2 = scraping_driver.find_element(By.XPATH, "/html/body/div[1]/div/div/div[2]/div[2]/form/button")
        elem2.click()
        time.sleep(10)
        scraping_driver.get(f"https://partner.ddnayo.com/reservationManagement/bookingRequestList?isStay=false&accommodationId={dn_no}")
        # deagi = driver.find_element(By.XPATH, "/html/body/div[1]/div/header/div/div[2]/div[5]/div[2]/a")
        # deagi.click()
        time.sleep(5)
        ui.system_status.append("scraping start.")
        scrape_data(scraping_driver)
        scraping_driver.quit()
    try:
        ui.system_status.append("스크래핑 작업이 진행 중입니다...\n")
        # 스크래핑 시작
        start_scraper()

        ui.system_status.append("스크래핑 작업이 완료되었습니다.\n")
    finally:
        pass


    
def generate_yearly_excel_files(year, room_names=None):
    """연도별로 Excel 파일을 생성하는 함수."""
    if room_names is None:
        room_names = ["A201", "A202", "B501", "B502", "B503", "B601", "B602", "B603", "사랑채", "C02", "C03", "C04", "C05", "별채"]

    columns = ["객실", "예약신청", "예약자", "이용기간", "결제일", "정산금액", "결제수단", "예약상태", "안원추가", "수영장옵션", "바베큐", "피크닉세트"]

    if getattr(sys, 'frozen', False):
        BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
    else:
        BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    folder_path = os.path.join(BASE_DIR, 'excel')
    os.makedirs(folder_path, exist_ok=True)

    def create_monthly_workbook(year, month):
        wb = Workbook()
        dates = pd.date_range(start=f'{year}-{month:02d}-01', end=f'{year}-{month:02d}-{pd.Timestamp(year=year, month=month, day=1).days_in_month}', freq='D')
        for date in dates:
            date_str = date.strftime("%Y-%m-%d")
            df = pd.DataFrame(room_names, columns=["객실"])
            for col in columns[1:]:
                df[col] = ""
            ws = wb.create_sheet(title=date_str)
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)
            column_widths = {'B': 28, 'C': 27.5, 'D': 23, 'E': 9, 'F': 13, 'G': 9, 'H': 9}
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                if column in column_widths:
                    ws.column_dimensions[column].width = column_widths[column]
                else:
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column].width = adjusted_width
            
            thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = thin_border
        
        del wb["Sheet"]
        return wb

    for month in range(1, 13):
        file_name = f"{year}-{month:02d}.xlsx"
        file_path = os.path.join(folder_path, file_name)
        if not os.path.exists(file_path):
            wb = create_monthly_workbook(year, month)
            wb.save(file_path)
            print(f"Saved {file_path}")
        else:
            print(f"Skipped {file_path} (already exists)")
     
