# -*- coding: utf-8 -*-
# Ilmare Pension ChatBot (rev. 2025‑05‑05)

# ── stdlib ────────────────────────────────────────────────────────────────
import os, sys, re, json, logging
from datetime import datetime, timedelta, date
from typing import List, Optional

# ── 3rd‑party ─────────────────────────────────────────────────────────────
from dotenv import load_dotenv
from openai import OpenAI
import openpyxl

# ── 경로 · 상수 설정 ──────────────────────────────────────────────────────
if getattr(sys, "frozen", False):                    # PyInstaller 대응
    BASE_DIR = os.path.dirname(os.path.abspath(sys.executable))
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

JSON_DIR   = os.path.join(BASE_DIR, "json")
EXCEL_DIR  = os.path.join(BASE_DIR, "excel")
UNANSWERED = os.path.join(JSON_DIR, "unanswered_questions.json")

os.makedirs(JSON_DIR,  exist_ok=True)
os.makedirs(EXCEL_DIR, exist_ok=True)
if not os.path.exists(UNANSWERED):
    with open(UNANSWERED, "w", encoding="utf‑8") as f:
        json.dump([], f, ensure_ascii=False, indent=4)

# ── 공통 유틸 ─────────────────────────────────────────────────────────────
SEOULOFFSET = timedelta(hours=9)

def today() -> date:
    """서울(UTC+9) 기준 오늘 날짜(date)."""
    return (datetime.utcnow() + SEOULOFFSET).date()

def save_unanswered_question(q: str) -> None:
    if not q.strip():
        return
    with open(UNANSWERED, "r", encoding="utf‑8") as f:
        data: List[str] = json.load(f)
    if q not in data:
        data.append(q)
        with open(UNANSWERED, "w", encoding="utf‑8") as f:
            json.dump(data, f, ensure_ascii=False, indent=4)

def load_room_names() -> List[str]:
    try:
        with open(os.path.join(JSON_DIR, "rooms.json"), encoding="utf‑8") as f:
            return json.load(f).get("rooms", [])
    except Exception:
        return []

# ── 자연어 날짜 파서 ──────────────────────────────────────────────────────
_REL = {"오늘": 0, "내일": 1, "모레": 2, "내일모레": 2, "어제": -1}

def parse_human_date(s: str, base: Optional[date] = None) -> Optional[date]:
    """'오늘', '7월 3일', '2025‑07‑03' → date 객체"""
    if not s:
        return None
    base = base or today()

    # 1) 상대 표현
    if s in _REL:
        return base + timedelta(days=_REL[s])

    # 2) YYYY‑MM‑DD
    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        pass

    # 3) '7월 3일'
    m = re.match(r"(\d{1,2})\s*월\s*(\d{1,2})\s*일?", s)
    if m:
        month, day = map(int, m.groups())
        y = base.year
        cand = date(y, month, day)
        if cand < base:                              # 이미 지났으면 다음 해
            cand = date(y + 1, month, day)
        return cand
    return None

# ═══════════════════════════════════════════════════════════════════════════
# 기능 1: 제부도 바닷길 통행 시간
# ═══════════════════════════════════════════════════════════════════════════
def get_tide_data(target_date: str) -> str:
    excel = os.path.join(EXCEL_DIR, "sea_road.xlsx")
    if not os.path.exists(excel):
        return "Error: sea_road.xlsx 파일을 찾을 수 없습니다."

    date_obj = parse_human_date(target_date)
    if not date_obj:
        return "Error: 날짜 형식이 잘못되었습니다. 'YYYY-MM-DD' 또는 오늘/내일/모레를 입력."

    wb = openpyxl.load_workbook(excel, read_only=True, data_only=True)
    sh = wb.active
    for row in sh.iter_rows(values_only=True):
        if isinstance(row[0], datetime) and row[0].date() == date_obj:
            t = [str(v).zfill(4) for v in row[3:7]]
            wb.close()
            return (
                f"{date_obj} 제부도 바닷길 통행 가능 시간\n"
                f"① {t[0][:2]}:{t[0][2:]} ~ {t[1][:2]}:{t[1][2:]}\n"
                f"② {t[2][:2]}:{t[2][2:]} ~ {t[3][:2]}:{t[3][2:]}"
            )
    wb.close()
    return "해당 날짜의 물때 정보를 찾을 수 없습니다."

# ═══════════════════════════════════════════════════════════════════════════
# 기능 2: 객실 예약 가능 여부
# ═══════════════════════════════════════════════════════════════════════════
def generate_reservation_link(checkin: str, checkout: str, room: str) -> str:
    mapping = {
        "201호": "5601454", "201": "5601454",
        "202호": "5601478", "202": "5601478",
        "501호": "5601501", "501": "5601501",
        "502호": "5601520", "502": "5601520",
        "503호": "5601527", "503": "5601527",
        "601호": "5601533", "601": "5601533",
        "602호": "5601537", "602": "5601537",
        "603호": "5601542", "603": "5601542",
        "사랑채": "5601549",
        "C02호": "5870382", "C02": "5870382", "c02": "5870382", "c02호": "5870382",
        "C03호": "5807525", "C03": "5807525", "c03": "5807525", "c03호": "5807525",
        "C04호": "5807529", "C04": "5807529", "co4": "5807529", "co4호": "5807529",
        "C05호": "5807541", "C05": "5807541", "co5": "5807541", "co5호": "5807541",
        "별채": "6369700",
    }
    room_no = mapping.get(room, "5601454")
    fmt = lambda d: datetime.strptime(d, "%Y%m%d").strftime("%Y년 %m월 %d일")
    return (
        f"입실: {fmt(checkin)}\n"
        f"퇴실: {fmt(checkout)}\n"
        f"객실: {room}\n"
        f"예약 링크: https://m.place.naver.com/accommodation/20530084/room/{room_no}"
        f"?checkin={checkin}&checkout={checkout}"
    )

def find_value_in_excel(month: int, day: int, target_value: Optional[str] = None,
                        year: Optional[int] = None) -> str:
    rooms = load_room_names()
    if not rooms:
        return "Error: 객실 목록을 로드할 수 없습니다."

    base = today()
    year = year or base.year
    try:
        query = date(year, month, day)
    except ValueError:
        return "Error: 잘못된 날짜입니다."

    if query < base:
        query = date(year + 1, month, day)

    y, m, d = query.year, query.month, query.day
    file   = os.path.join(EXCEL_DIR, f"{y}-{m:02}.xlsx")
    sheet  = f"{y}-{m:02}-{d:02}"

    if not os.path.exists(file):
        return f"Error: {file} 파일이 없습니다."

    try:
        wb = openpyxl.load_workbook(file, data_only=True)
        if sheet not in wb.sheetnames:
            wb.close()
            return f"Error: '{sheet}' 시트를 찾을 수 없습니다."

        sh = wb[sheet]

        def is_reserved(room: str) -> bool:
            for r in sh.iter_rows(min_col=1, max_col=3, values_only=True):
                if room in str(r[0]) and r[2]:
                    return True
            return False

        # 방 번호를 지정하지 않은 경우
        if not target_value:
            wb.close()
            return "조회하려는 객실을 지정해 주세요. (예: 201호)"

        if is_reserved(target_value):
            pools, generals, groups = [], [], []
            for rm in rooms:
                if not is_reserved(rm):
                    if rm in ["A201", "A202", "B501", "B502", "B503", "B601", "B602", "B603"]:
                        pools.append(rm)
                    elif rm == "별채":
                        groups.append(rm)
                    else:
                        generals.append(rm)
            wb.close()
            return (f"{query} '{target_value}'는 이미 예약되었습니다.\n"
                    f"가능 객실\n • 풀빌라: {', '.join(pools) or '없음'}\n"
                    f" • 일반: {', '.join(generals) or '없음'}\n"
                    f" • 단체: {', '.join(groups) or '없음'}")
        # 예약 가능
        checkout = (datetime.combine(query, datetime.min.time()) + timedelta(days=1)).strftime("%Y%m%d")
        link_msg = generate_reservation_link(query.strftime("%Y%m%d"), checkout, target_value)
        wb.close()
        return f"{query} '{target_value}' 예약 가능합니다.\n{link_msg}"
    except Exception as e:
        logging.exception(e)
        return f"Error: {e}"

# ── OpenAI 툴 스키마 (year, target_value 선택 사항) ─────────────────────────
tools = [
    {
        "type": "function",
        "function": {
            "name": "find_value_in_excel",
            "description": "연/월/일/객실명으로 예약 가능 여부 조회",
            "parameters": {
                "type": "object",
                "properties": {
                    "year":  {"type":"string","description":"생략 가능"},
                    "month": {"type":"string","description":"1~12"},
                    "day":   {"type":"string","description":"1~31"},
                    "target_value":{"type":"string","description":"객실명(201호 등)"},
                },
                "required": ["month", "day"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_tide_data",
            "description": "제부도 바닷길(물때) 통행 가능 시간 조회",
            "parameters": {
                "type":"object",
                "properties":{
                    "target_date":{"type":"string","description":"YYYY-MM-DD or 오늘/내일"}
                },
                "required":["target_date"]
            }
        }
    }
]

# ═══════════════════════════════════════════════════════════════════════════
# ChatBot 클래스 (2‑스텝 툴 호출)
# ═══════════════════════════════════════════════════════════════════════════
class ChatBot:
    def __init__(self, model: str, system_prompt: str):
        load_dotenv()
        self.client   = OpenAI()
        self.model    = model
        self.messages = [{"role":"system","content":system_prompt}]

    def add(self, role: str, content: str) -> None:
        self.messages.append({"role":role,"content":content or "내용 없음"})

    # ───────────────────────────────────────────────────────────────────────
    def _call_tool(self, name: str, args: dict) -> str:
        if name == "find_value_in_excel":
            y = int(args["year"]) if args.get("year") else None
            return find_value_in_excel(month=int(args["month"]),
                                       day=int(args["day"]),
                                       target_value=args.get("target_value"),
                                       year=y)
        elif name == "get_tide_data":
            return get_tide_data(args["target_date"])
        return "Error: Unknown tool."

    # ───────────────────────────────────────────────────────────────────────
    def get_response(self, user_input: str) -> str:
        if not user_input.strip():
            return ""
        self.add("user", user_input)

        # 1) 모델에게 툴 호출 판단 요청
        first = self.client.chat.completions.create(
            model=self.model, messages=self.messages, tools=tools
        )
        msg = first.choices[0].message

        # 툴 호출이 없다면 바로 답변
        if not msg.tool_calls:
            answer = msg.content or ""
            self.add("assistant", answer)
            if not answer.strip():
                save_unanswered_question(user_input)
                answer = "[AI 자동답변] : 죄송합니다. 이 질문에 대한 답변을 준비하지 못했습니다."
            return answer

        # 2) 툴 실행
        tc = msg.tool_calls[0]
        args = json.loads(tc.function.arguments)
        tool_result = self._call_tool(tc.function.name, args)

        # 3) tool 역할 메시지 추가 후 재호출
        self.messages.append({
            "role": "tool",
            "tool_call_id": tc.id,
            "content": tool_result
        })

        second = self.client.chat.completions.create(
            model=self.model, messages=self.messages
        )
        answer = second.choices[0].message.content or ""
        self.add("assistant", answer)

        if not answer.strip():
            save_unanswered_question(user_input)
            answer = "[AI 자동답변] : 죄송합니다. 이 질문에 대한 답변을 준비하지 못했습니다."
        return answer

    # ───────────────────────────────────────────────────────────────────────
    def reset(self) -> None:
        """시스템 프롬프트만 남기고 초기화."""
        self.messages = self.messages[:1]
