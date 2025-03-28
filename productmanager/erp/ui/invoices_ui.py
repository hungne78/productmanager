from PyQt5.QtWidgets import QWidget, QVBoxLayout, QHBoxLayout, QPushButton, QTableWidget, QTableWidgetItem, \
    QHeaderView, QLabel, QComboBox, QFileDialog, QLineEdit, QGroupBox
import requests
import pandas as pd
from PyQt5.QtCore import Qt, QSize
import sys
import os
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))
from services.api_services import get_auth_headers
from PyQt5.QtWidgets import QSizePolicy

BASE_URL = "http://127.0.0.1:8000"  # 실제 서버 URL
global_token = get_auth_headers  # 로그인 토큰 (Bearer 인증)
import pandas as pd
from datetime import datetime


class InvoicesLeftPanel(QWidget):
    """
    왼쪽 패널 - 연도 & 월 선택 + 유형 선택 + 등록/수정/삭제 버튼
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self.parent_widget = parent
        self.init_ui()

    def init_ui(self):
        main_layout = QVBoxLayout()

        # ✅ [1] 연도 & 월 선택 (그룹박스로 감싸기)
        date_group = QGroupBox("📅 기간 선택")
        date_layout = QHBoxLayout()
        self.year_selector = QComboBox()
        self.month_selector = QComboBox()

        # 📌 크기 조정
        self.year_selector.setMinimumSize(QSize(80, 30))
        self.month_selector.setMinimumSize(QSize(80, 30))

        # 📅 현재 연도 & 월 설정
        current_year = datetime.today().year
        current_month = datetime.today().month
        for y in range(current_year - 5, current_year + 6):
            self.year_selector.addItem(str(y))
        for m in range(1, 13):
            self.month_selector.addItem(str(m).zfill(2))

        self.year_selector.setCurrentText(str(current_year))
        self.month_selector.setCurrentText(str(current_month).zfill(2))

        # 📅 UI 배치
        date_layout.addWidget(QLabel("연도:"))
        date_layout.addWidget(self.year_selector)
        date_layout.addWidget(QLabel("월:"))
        date_layout.addWidget(self.month_selector)
        date_group.setLayout(date_layout)
        main_layout.addWidget(date_group)

        # ✅ [2] 세금계산서 유형 선택 (그룹박스로 감싸기)
        type_group = QGroupBox("📑 세금계산서 유형")
        type_layout = QVBoxLayout()
        self.invoice_type_selector = QComboBox()
        self.invoice_type_selector.addItem("01 (일반)")
        self.invoice_type_selector.addItem("02 (영세율)")
        self.invoice_type_selector.setMinimumSize(QSize(140, 30))

        type_layout.addWidget(self.invoice_type_selector)
        type_group.setLayout(type_layout)
        main_layout.addWidget(type_group)

        # ✅ [3] 거래처 검색 필터
        search_group = QGroupBox("🔍 거래처 검색")
        search_layout = QVBoxLayout()
        self.client_search = QLineEdit()
        self.client_search.setPlaceholderText("거래처명 입력...")
        self.client_search.setMinimumSize(QSize(200, 30))
        self.client_search.textChanged.connect(self.filter_clients)

        search_layout.addWidget(self.client_search)
        search_group.setLayout(search_layout)
        main_layout.addWidget(search_group)

        # ✅ [4] 총 공급가액 & 총 세액 (가운데 정렬)
        total_group = QGroupBox("💰 총 금액")
        total_layout = QVBoxLayout()
        self.total_sales_label = QLabel("총 공급가액: ₩0")
        self.total_tax_label = QLabel("총 세액: ₩0")

        # 📌 가운데 정렬
        self.total_sales_label.setAlignment(Qt.AlignCenter)
        self.total_tax_label.setAlignment(Qt.AlignCenter)

        total_layout.addWidget(self.total_sales_label)
        total_layout.addWidget(self.total_tax_label)
        total_group.setLayout(total_layout)
        main_layout.addWidget(total_group)

        # ✅ [5] 버튼 (등록/수정/삭제) 세로 정렬
        btn_group = QGroupBox("🛠️ 관리")
        btn_layout = QVBoxLayout()
        self.add_button = QPushButton("➕ 등록")
        self.edit_button = QPushButton("✏️ 수정")
        self.delete_button = QPushButton("🗑 삭제")

        # 📌 버튼 크기 조정
        self.add_button.setMinimumSize(QSize(120, 40))
        self.edit_button.setMinimumSize(QSize(120, 40))
        self.delete_button.setMinimumSize(QSize(120, 40))
        self.add_button.clicked.connect(self.on_add_clicked)
        self.edit_button.clicked.connect(self.on_edit_clicked)
        self.delete_button.clicked.connect(self.on_delete_clicked)
        btn_layout.addWidget(self.add_button)
        btn_layout.addWidget(self.edit_button)
        btn_layout.addWidget(self.delete_button)
        btn_group.setLayout(btn_layout)
        main_layout.addWidget(btn_group)

        # ✅ [6] 📊 조회 버튼 (가로로 넓게 배치)
        self.search_button = QPushButton("📊 세금계산서 조회")
        self.search_button.setMinimumSize(QSize(250, 40))
        self.search_button.setStyleSheet("font-size: 14px; font-weight: bold;")
        main_layout.addWidget(self.search_button)

        # ✅ [7] 📂 파일 불러오기 버튼 (하단)
        self.import_button = QPushButton("📂 파일 불러오기")
        self.import_button.setMinimumSize(QSize(250, 40))
        self.import_button.clicked.connect(self.import_excel)
        main_layout.addWidget(self.import_button)

        # 📌 레이아웃 설정
        self.search_button.clicked.connect(self.fetch_invoices)
        self.setLayout(main_layout)

    def fetch_invoices(self):
        """
        세금계산서 조회
        """
        if self.parent_widget:
            selected_year = self.year_selector.currentText()
            selected_month = self.month_selector.currentText()
            self.parent_widget.load_invoices(selected_year, selected_month)

    def filter_clients(self):
        """
        거래처명을 입력하면 해당 거래처만 필터링하여 조회
        """
        if self.parent_widget:
            search_text = self.client_search.text()
            self.parent_widget.filter_invoices(search_text)

    def update_totals(self, total_sales, total_tax):
        """
        총 공급가액 & 총 세액 표시 업데이트
        """
        self.total_sales_label.setText(f"💰 총 공급가액: ₩{total_sales:,}")
        self.total_tax_label.setText(f"💵 총 세액: ₩{total_tax:,}")

    def import_excel(self):
        """
        엑셀 파일 불러오기
        """
        file_path, _ = QFileDialog.getOpenFileName(self, "엑셀 파일 불러오기", "", "Excel Files (*.xls *.xlsx)")
        if not file_path:
            return

        try:
            df = pd.read_excel(file_path, sheet_name="Sheet1", skiprows=3)  # 헤더 제외하고 데이터 로드
            invoice_data = df.to_dict(orient="records")
            self.parent_widget.load_imported_invoices(invoice_data)
            print(f"✅ 엑셀 파일 로드 성공: {file_path}")
        except Exception as e:
            print(f"❌ 엑셀 파일 로드 실패: {e}")
    def on_add_clicked(self):
        """
        등록 버튼 눌렀을 때 -> 작은 다이얼로그 통해 새 invoice 데이터 입력받고,
        parent_widget.add_invoice(...)로 추가
        """
        from PyQt5.QtWidgets import QDialog, QFormLayout, QLineEdit, QDialogButtonBox

        class AddInvoiceDialog(QDialog):
            def __init__(self, parent=None):
                super().__init__(parent)
                self.setWindowTitle("새 세금계산서 등록")

                self.client_id_edit = QLineEdit()
                self.client_name_edit = QLineEdit()
                self.client_ceo_edit = QLineEdit()
                self.sales_edit = QLineEdit()
                self.tax_edit = QLineEdit()

                form = QFormLayout()
                form.addRow("거래처 ID:", self.client_id_edit)
                form.addRow("거래처명:", self.client_name_edit)
                form.addRow("대표자명:", self.client_ceo_edit)
                form.addRow("공급가액:", self.sales_edit)
                form.addRow("세액:", self.tax_edit)

                btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                btn_box.accepted.connect(self.accept)
                btn_box.rejected.connect(self.reject)

                form.addWidget(btn_box)
                self.setLayout(form)

            def get_data(self):
                return {
                    "client_id": self.client_id_edit.text().strip(),
                    "client_name": self.client_name_edit.text().strip(),
                    "client_ceo": self.client_ceo_edit.text().strip(),
                    "total_sales": float(self.sales_edit.text() or 0),
                    "tax_amount": float(self.tax_edit.text() or 0),
                }

        dialog = AddInvoiceDialog(self)
        if dialog.exec_() == QDialog.Accepted:
            data = dialog.get_data()
            self.parent_widget.add_invoice(data)  # InvoicesTab.add_invoice() 호출

    def on_edit_clicked(self):
        """
        수정 버튼 -> 우측 테이블에서 선택된 행 찾아서 수정
        """
        # (1) 우측 패널의 invoice_table에서 현재 선택된 행 인덱스 가져오기
        table = self.parent_widget.right_panel.invoice_table
        current_row = table.currentRow()
        if current_row < 0:
            print("⚠️ 수정할 행을 선택하세요.")
            return

        # (2) 현재행의 데이터(=parent_widget.all_invoices[current_row])를 참고해서
        #     대화창에 표시
        old_data = {}
        if 0 <= current_row < len(self.parent_widget.all_invoices):
            old_data = self.parent_widget.all_invoices[current_row]
        else:
            print("⚠️ all_invoices 범위 벗어남.")
            return

        from PyQt5.QtWidgets import QDialog, QFormLayout, QLineEdit, QDialogButtonBox

        class EditInvoiceDialog(QDialog):
            def __init__(self, parent=None, data=None):
                super().__init__(parent)
                self.setWindowTitle("세금계산서 수정")

                self.type_selector = QComboBox()
                self.type_selector.addItems(["01", "02"])
                self.type_selector.setCurrentText(data.get("type", "01"))

                self.client_id_edit = QLineEdit(data.get("client_id", ""))
                self.client_name_edit = QLineEdit(data.get("client_name", ""))
                self.client_ceo_edit = QLineEdit(data.get("client_ceo", ""))
                self.sales_edit = QLineEdit(str(data.get("total_sales", 0)))
                self.tax_edit = QLineEdit(str(data.get("tax_amount", 0)))
                # ✅ 새 항목 추가: 영수/청구
                self.rc_selector = QComboBox()
                self.rc_selector.addItems(["01", "02"])
                self.rc_selector.setCurrentText(data.get("rc_type", "01"))

                form = QFormLayout()
                form.addRow("종류 (01/02):", self.type_selector)  # ✅ 추가
                form.addRow("거래처 ID:", self.client_id_edit)
                form.addRow("거래처명:", self.client_name_edit)
                form.addRow("대표자명:", self.client_ceo_edit)
                form.addRow("공급가액:", self.sales_edit)
                form.addRow("세액:", self.tax_edit)
                form.addRow("영수/청구:", self.rc_selector)

                btn_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
                btn_box.accepted.connect(self.accept)
                btn_box.rejected.connect(self.reject)

                form.addWidget(btn_box)
                self.setLayout(form)

            def get_data(self):
                return {
                    "type": self.type_selector.currentText(),               # ✅ 추가
                    "client_id": self.client_id_edit.text().strip(),
                    "client_name": self.client_name_edit.text().strip(),
                    "client_ceo": self.client_ceo_edit.text().strip(),
                    "total_sales": float(self.sales_edit.text() or 0),
                    "tax_amount": float(self.tax_edit.text() or 0),
                    "rc_type": self.rc_selector.currentText(),              # ✅ 추가
                }

        dialog = EditInvoiceDialog(self, data=old_data)
        if dialog.exec_() == QDialog.Accepted:
            new_data = dialog.get_data()
            # InvoicesTab.update_invoice(행인덱스, 수정된딕셔너리)
            self.parent_widget.update_invoice(current_row, new_data)

    def on_delete_clicked(self):
        """
        삭제 버튼 -> 우측 테이블에서 선택된 행 삭제
        """
        table = self.parent_widget.right_panel.invoice_table
        current_row = table.currentRow()
        if current_row < 0:
            print("⚠️ 삭제할 행을 선택하세요.")
            return

        self.parent_widget.delete_invoice(current_row)
class InvoicesRightPanel(QWidget):
    def __init__(self):
        super().__init__()
        self.company_info = {}
        self.init_ui()

    def set_company_info(self, info: dict):
        self.company_info = info or {}

        company_name = self.company_info.get("company_name")
        ceo_name = self.company_info.get("ceo_name") or self.company_info.get("ceo", "")
        business_number = self.company_info.get("business_number")

        if company_name and ceo_name and business_number:
            txt = f"[{company_name}] 대표: {ceo_name} / 사업자번호: {business_number}"
            self.company_label.setText(txt)
        else:
            self.company_label.setText("공급자(우리 회사) 정보가 등록되지 않았습니다.")



    def init_ui(self):
        layout = QVBoxLayout()

        # (1) 상단 라벨에 회사 정보 표시(간단 예시)
        self.company_label = QLabel("공급자(우리 회사) 정보가 등록되지 않았습니다.")
        layout.addWidget(self.company_label)

        self.invoice_table = QTableWidget()
        self.invoice_table.setColumnCount(14)
        self.invoice_table.setHorizontalHeaderLabels([
           "종류",              # (A)
        "작성일자",          # (B)
        "공급자 등록번호",    # (C)
        "공급자 상호",       # (E)
        "공급자 성명",       # (F)
        "공급받는자 등록번호",# (K)
        "공급받는자 상호",   # (M)
        "공급받는자",        # (N)
        "공급가액 합계",     # (T)
        "세액 합계",         # (U)
        "일자1",            # (W)
        "공급가액1",         # (AB)
        "세액1",            # (AC)
        "영수/청구",         # (BG) → 기본값 "01"
        ])

        # ── 컬럼 리사이즈 옵션 ───────────────────────────
        # 1) 모든 열을 내용에 맞춰 크기 조정
        self.invoice_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        # 또는 일부 열은 늘이고, 일부 열은 고정 너비 등 원하는 대로 커스텀 가능

        layout.addWidget(QLabel("📑 거래처별 월 매출 목록"))
        layout.addWidget(self.invoice_table)

        # (2) 엑셀 저장 버튼
        self.export_button = QPushButton("📄 엑셀 저장")
        self.export_button.clicked.connect(self.export_to_excel)
        layout.addWidget(self.export_button)

        self.setLayout(layout)

    def update_invoice_data(self, invoice_data):
        # 회사정보 라벨 세팅(생략)
        ...
        self.invoice_table.setRowCount(0)
        today_day = datetime.today().strftime("%d") 
        for invoice in invoice_data:
            row = self.invoice_table.rowCount()
            self.invoice_table.insertRow(row)

            # 🔹 필요 변수들
            supplier_reg  = self.company_info.get("business_number", "")
            supplier_name = self.company_info.get("company_name", "")
            supplier_ceo  = self.company_info.get("ceo_name") or self.company_info.get("ceo", "")
            
            client_reg  = invoice.get("business_number", "")
            client_name = invoice.get("client_name", "")
            client_ceo  = invoice.get("client_ceo", "")

            total_sales = invoice["total_sales"]
            tax_amount  = invoice["tax_amount"]

            # ✅ 추가: 수정 다이얼로그에서 받아온 값 처리
            doc_type = invoice.get("type", "01")       # 기본 "01"
            rc_type  = invoice.get("rc_type", "01")    # 기본 "01"

            # 🔹 테이블 셀 채우기
            self.invoice_table.setItem(row, 0, QTableWidgetItem(doc_type))  # ✅ 종류
            self.invoice_table.setItem(row, 1, QTableWidgetItem(datetime.today().strftime("%Y%m%d")))  # 작성일자
            self.invoice_table.setItem(row, 2, QTableWidgetItem(supplier_reg))
            self.invoice_table.setItem(row, 3, QTableWidgetItem(supplier_name))
            self.invoice_table.setItem(row, 4, QTableWidgetItem(supplier_ceo))

            self.invoice_table.setItem(row, 5, QTableWidgetItem(client_reg))
            self.invoice_table.setItem(row, 6, QTableWidgetItem(client_name))
            self.invoice_table.setItem(row, 7, QTableWidgetItem(client_ceo))

            self.invoice_table.setItem(row, 8, QTableWidgetItem(f"{int(total_sales):,}"))
            self.invoice_table.setItem(row, 9, QTableWidgetItem(f"{int(tax_amount):,}"))

            self.invoice_table.setItem(row, 10, QTableWidgetItem(today_day))          # ✅ 오늘 일자(2자리)
            self.invoice_table.setItem(row, 11, QTableWidgetItem(f"{int(total_sales):,}"))
            self.invoice_table.setItem(row, 12, QTableWidgetItem(f"{int(tax_amount):,}"))

            self.invoice_table.setItem(row, 13, QTableWidgetItem(rc_type))   

    def export_to_excel(self):
        import openpyxl
        from openpyxl import load_workbook
        from PyQt5.QtWidgets import QFileDialog
        import shutil

        # 1) 사용자에게 저장경로 묻기
        save_path, _ = QFileDialog.getSaveFileName(
            self,
            "엑셀 파일 저장",
            "",
            "Excel Files (*.xlsx *.xls)"  # 가능하다면 xlsx를 권장
        )
        if not save_path:
            return

        # 2) 테이블 데이터 수집
        row_count = self.invoice_table.rowCount()
        col_count = self.invoice_table.columnCount()

        import os

        base_dir = os.path.dirname(os.path.abspath(__file__))  # invoices_ui.py가 있는 폴더
        template_dir = os.path.join(base_dir, "templates")

        if row_count <= 100:
            template_path = os.path.join(template_dir, "세금계산서등록양식(일반).xlsx")
        else:
            template_path = os.path.join(template_dir, "세금계산서등록양식(일반)_대량.xlsx")
                # 3) 원본 템플릿 복사 → 임시 파일로 (xls인 경우 openpyxl이 호환 문제 있을 수도 있으니 xlsx 권장)
        temp_copy = "temp_copy.xlsx"
        shutil.copyfile(template_path, temp_copy)

        # 4) openpyxl로 임시 파일 열기 (xls는 지원이 제한되므로 실제론 xlsx를 권장합니다)
        try:
            wb = load_workbook(temp_copy)
        except Exception as e:
            print(f"❌ 템플릿 열기 실패: {e}")
            return

        ws = wb.active  # "Sheet1" 가정

        # 5) 엑셀에서 7행부터 채우고 싶다면
        start_row = 7

        # 6) 열 매핑 딕셔너리
        #    Key: 테이블 헤더이름, Value: 엑셀 열 주소
        column_mapping = {
            "종류":           "A",
            "작성일자":       "B",
            "공급자 등록번호": "C",
            "공급자 상호":    "E",
            "공급자 성명":    "F",
            "공급받는자 등록번호": "K",
            "공급받는자 상호":   "M",
            "공급받는자":       "N",
            "공급가액 합계":     "T",
            "세액 합계":       "U",
            "일자1":          "W",
            "공급가액1":       "AB",
            "세액1":          "AC",
            "영수/청구":       "BG",  # 기본값 "01"
        }

        # 7) 실제 테이블로부터 셀 값을 읽어, 매핑된 열에 기입
        for r in range(row_count):
            excel_row = start_row + r

            for c in range(col_count):
                header_text = self.invoice_table.horizontalHeaderItem(c).text()  # "종류", "작성일자" 등
                cell_widget = self.invoice_table.item(r, c)
                if cell_widget:
                    val = cell_widget.text().strip()
                else:
                    val = ""

                # 영수/청구 열 → 빈값이면 "01"로
                if header_text == "영수/청구" and not val:
                    val = "01"

                if header_text in column_mapping:
                    excel_col = column_mapping[header_text]  # 예: "A"
                    target_cell = f"{excel_col}{excel_row}"
                    ws[target_cell] = val
                else:
                    # 매핑에 없는 열은 무시
                    pass

        # 8) 저장
        try:
            wb.save(temp_copy)
            shutil.copyfile(temp_copy, save_path)  # 최종 저장
            print(f"✅ 엑셀 저장 성공: {save_path}")
        except Exception as e:
            print(f"❌ 엑셀 저장 실패: {e}")
        finally:
            # 임시파일 정리 등
            pass


class InvoicesTab(QWidget):
    def __init__(self):
        super().__init__()
        layout = QHBoxLayout()
        self.left_panel = InvoicesLeftPanel(self)
        self.right_panel = InvoicesRightPanel()
        
        # ✅ 모든 조회 결과를 저장해둘 리스트
        self.all_invoices = []
        
        # ✅ 크기 정책 설정
        self.left_panel.setSizePolicy(QSizePolicy.Fixed, QSizePolicy.Expanding)
        self.right_panel.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)

        # ✅ 고정 크기 설정
        self.left_panel.setFixedWidth(350)  # 1 비율
        layout.addWidget(self.left_panel)
        layout.addWidget(self.right_panel)
        self.setLayout(layout)
        self.fetch_company_info()
        self.setStyleSheet("""
QWidget {
    background-color: #F7F9FC; /* 좀 더 밝은 배경 */
    font-family: 'Malgun Gothic', 'Segoe UI', sans-serif;
    color: #2F3A66;
}
QGroupBox {
    background-color: rgba(255,255,255, 0.8);
    border: 1px solid #E2E8F0;
    border-radius: 12px;
    padding: 16px;
    margin-top: 12px;
}
QGroupBox::title {
    font-size: 15px;
    font-weight: 600;
    color: #4B5D88;
    padding: 6px 12px;
}
QPushButton {
    background-color: #E2E8F0;
    border: 2px solid #CBD5E0;
    border-radius: 6px;
    padding: 8px 14px;
    font-weight: 500;
    color: #2F3A66;
}
QPushButton:hover {
    background-color: #CBD5E0;
}
QTableWidget {
    background-color: #FFFFFF;
    border: 3px solid #D2D6DC;
    border-radius: 8px;
    gridline-color: #E2E2E2;
    font-size: 15px;
    color: #333333;
    alternate-background-color: #fafafa;
    selection-background-color: #c8dafc;
}
QHeaderView::section {
    background-color: #EEF1F5;
    color: #333333;
    font-weight: 600;
    padding: 6px;
    border: 1px solid #D2D6DC;
    border-radius: 0;
    border-bottom: 2px solid #ddd;
}
""")
    def fetch_company_info(self):
        """ 서버에서 회사 정보를 가져와 우측 패널에 표시 """
        url = f"http://127.0.0.1:8000/company"
        headers = {"Authorization": f"Bearer {global_token}"}

        try:
            resp = requests.get(url, headers=headers)
            resp.raise_for_status()
            company_info = resp.json()

            # ✅ 우측 패널에 회사 정보 업데이트
            self.right_panel.set_company_info(company_info)

        except Exception as e:
            print(f"❌ 회사 정보 조회 실패: {e}")

    def load_invoices(self, year, month):
        """ 거래처별 월 매출 데이터 로드 """
        url = f"http://127.0.0.1:8000/sales/clients/{year}/{month}"
        headers = {"Authorization": f"Bearer {global_token}"}

        try:
            resp = requests.get(url, headers=headers)
            resp.raise_for_status()
            invoice_data = resp.json()

            # ✅ 세액 계산 방식 변경: 총 매출의 90% / 10% 분리
            for invoice in invoice_data:
                total = float(invoice["total_sales"])
                invoice["total_sales"] = round(total * 0.9)  # ✅ 90% 설정
                invoice["tax_amount"] = round(total * 0.1)  # ✅ 10% 설정

            self.all_invoices = invoice_data
            self.right_panel.update_invoice_data(self.all_invoices)

        except Exception as e:
            print(f"❌ 거래처 매출 조회 실패: {e}")
            self.all_invoices = []
            self.right_panel.update_invoice_data([])

    def filter_invoices(self, search_text: str):
        """ 거래처명 필터링 후 표시 """
        if not search_text:
            filtered = self.all_invoices
        else:
            filtered = [item for item in self.all_invoices if search_text.lower() in item.get("client_name", "").lower()]
        self.right_panel.update_invoice_data(filtered)
        
    def do_search(self, keyword: str):
        """
        메인 윈도우의 검색창과 연동되는 메서드
        - InvoicesTab을 보여준 상태에서, 검색창에 "거래처명"을 입력해 엔터/검색 버튼 누르면 여기로 옴.
        - 결국 filter_invoices와 동일한 동작
        """
        self.filter_invoices(keyword)
    
    def add_invoice(self, invoice_data: dict):
        """
        새 매출(세금계산서) 항목을 로컬 리스트에 추가
        invoice_data 예시:
          {
            "client_id": "1234",
            "client_name": "거래처A",
            "client_ceo": "홍길동",
            "total_sales": 50000,
            "tax_amount": 5000
          }
        """
        self.all_invoices.append(invoice_data)
        self.right_panel.update_invoice_data(self.all_invoices)

    def update_invoice(self, row_index: int, new_data: dict):
        """
        우측 테이블에서 선택된 row_index에 대해, new_data로 수정
        """
        if 0 <= row_index < len(self.all_invoices):
            self.all_invoices[row_index].update(new_data)
            # 혹은 완전히 대체하려면: self.all_invoices[row_index] = new_data
            self.right_panel.update_invoice_data(self.all_invoices)
        else:
            print("⚠️ 잘못된 row_index:", row_index)

    def delete_invoice(self, row_index: int):
        """
        우측 테이블에서 선택된 row_index 항목을 self.all_invoices에서 제거
        """
        if 0 <= row_index < len(self.all_invoices):
            del self.all_invoices[row_index]
            self.right_panel.update_invoice_data(self.all_invoices)
        else:
            print("⚠️ 잘못된 row_index:", row_index)
